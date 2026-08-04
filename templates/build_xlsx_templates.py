"""Builds the Inventory bulk-import and BOM upload .xlsx templates.

Column headers here must match the aliases in
backend/app/services/bulk_import.py (_FIELD_ALIASES) and
backend/app/services/bom_matcher.py (_NAME_ALIASES / _QTY_ALIASES)
exactly enough to auto-map — using the canonical header text (the first
alias in each list) guarantees a clean auto-match with no manual mapping
step needed on upload.
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

FONT_NAME = "Calibri"
HEADER_FILL = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
EXAMPLE_FONT = Font(name=FONT_NAME, italic=True, color="94A3B8", size=11)
BODY_FONT = Font(name=FONT_NAME, size=11)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=16, color="0F172A")
SUBTITLE_FONT = Font(name=FONT_NAME, size=11, color="475569")
SECTION_FONT = Font(name=FONT_NAME, bold=True, size=12, color="0F172A")
WRAP = Alignment(wrap_text=True, vertical="top")


def _style_instructions_sheet(ws, title, intro_lines, steps, notes):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 100

    row = 1
    ws.cell(row=row, column=1, value=title).font = TITLE_FONT
    row += 2

    for line in intro_lines:
        c = ws.cell(row=row, column=1, value=line)
        c.font = SUBTITLE_FONT
        c.alignment = WRAP
        row += 1
    row += 1

    c = ws.cell(row=row, column=1, value="How to use this template")
    c.font = SECTION_FONT
    row += 1
    for i, step in enumerate(steps, start=1):
        c = ws.cell(row=row, column=1, value=f"{i}. {step}")
        c.font = BODY_FONT
        c.alignment = WRAP
        row += 1
    row += 1

    if notes:
        c = ws.cell(row=row, column=1, value="Notes")
        c.font = SECTION_FONT
        row += 1
        for note in notes:
            c = ws.cell(row=row, column=1, value=f"• {note}")
            c.font = BODY_FONT
            c.alignment = WRAP
            row += 1


def _style_data_sheet(ws, headers, example_rows, col_widths):
    ws.sheet_view.showGridLines = False
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths[col_idx - 1]
    ws.freeze_panes = "A2"

    for r, example_row in enumerate(example_rows, start=2):
        for col_idx, value in enumerate(example_row, start=1):
            cell = ws.cell(row=r, column=col_idx, value=value)
            cell.font = EXAMPLE_FONT


def build_inventory_template(path):
    wb = Workbook()
    instructions = wb.active
    instructions.title = "Instructions"
    _style_instructions_sheet(
        instructions,
        title="Proforce Airsystems — Inventory Bulk Import Template",
        intro_lines=[
            "Use this to add new components or top up existing ones in bulk, instead of",
            "entering them one at a time. Fill in the 'Inventory Import' tab and upload it",
            "from Inventory → Bulk Import.",
        ],
        steps=[
            "Delete the example row (row 2, shown in grey italics) before filling in your own data.",
            "Fill in one row per component. Only 'Name' is required — every other column is optional.",
            "Save the file as .xlsx or .csv.",
            "Go to Inventory → Bulk Import, upload the file, and confirm the column mapping "
            "(it should auto-match since the headers here match the system's field names).",
            "Review the import summary: new parts are created, and rows that match an existing "
            "component by name/brand have their quantity added to the existing stock count "
            "rather than creating a duplicate.",
        ],
        notes=[
            "Quantity should be a whole number (received/added stock, not a replacement count).",
            "Category must match an existing category name exactly to be linked automatically; "
            "otherwise the component is created without a category (nothing is lost — assign "
            "it afterwards in Inventory).",
            "Image URL is optional — leave blank if you don't have one; you can attach a photo "
            "later from the component's detail view.",
            "Don't include a SKU column — internal SKUs are assigned by the system, not "
            "imported from vendor sheets.",
        ],
    )

    data = wb.create_sheet("Inventory Import")
    headers = ["Name", "Type", "Category", "Brand", "Description", "Quantity", "Image URL"]
    example_rows = [
        [
            "M3x10 Socket Head Screw",
            "Fastener",
            "Hardware",
            "Acme",
            "Stainless steel, DIN 912",
            "250",
            "",
        ]
    ]
    col_widths = [30, 16, 16, 16, 36, 12, 30]
    _style_data_sheet(data, headers, example_rows, col_widths)

    wb.active = 1
    wb.save(path)


def build_bom_template(path):
    wb = Workbook()
    instructions = wb.active
    instructions.title = "Instructions"
    _style_instructions_sheet(
        instructions,
        title="Proforce Airsystems — BOM Upload Template",
        intro_lines=[
            "Use this to check a project's Bill of Materials against current inventory stock.",
            "Fill in the 'BOM' tab and upload it from Inventory → Check BOM.",
        ],
        steps=[
            "Delete the example row (row 2, shown in grey italics) before filling in your own data.",
            "List one part per row, with the quantity your project needs.",
            "Part names don't need to match inventory exactly — the checker fuzzy-matches "
            "them against existing components (e.g. 'ESC 30A' will match 'ESC — 30A Blheli_S').",
            "Save the file as .xlsx or .csv and upload it from Inventory → Check BOM.",
            "Review the results: each line comes back Available, Low Stock, or Missing, with a "
            "suggested substitute for anything not found.",
        ],
        notes=[
            "This template is for checking/reserving stock against a project — to add new "
            "parts to inventory itself, use the separate Inventory Bulk Import template instead.",
            "Quantity defaults to 1 if left blank.",
        ],
    )

    data = wb.create_sheet("BOM")
    headers = ["Part Name", "Quantity"]
    example_rows = [["ESC 30A Blheli_S", "4"]]
    col_widths = [40, 14]
    _style_data_sheet(data, headers, example_rows, col_widths)

    wb.active = 1
    wb.save(path)


if __name__ == "__main__":
    build_inventory_template("Inventory_Bulk_Import_Template.xlsx")
    build_bom_template("BOM_Upload_Template.xlsx")
    print("done")
